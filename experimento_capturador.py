# LIBRERIAS
import openpyxl
import os

# VARIABLES GLOBALES
num_semana = 8
ruta_proyecto = '/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/'
nombre_archivo = f'SEMANA_0{num_semana}_REPORTE.xlsx'

ruta_archivo_origen = os.path.join(ruta_proyecto, f'SEMANA_0{num_semana}', nombre_archivo)
ruta_archivo_destino = os.path.join(ruta_proyecto, f'SEMANA_0{num_semana}')

# GENERAR LOS DATOS DE CAPTURA
# Cargamos el archivo de Excel con openpyxl
libro = openpyxl.load_workbook(ruta_archivo_origen)
hoja = libro.active

# Definimos la lista donde vamos a guardar los datos
datos_de_captura = []

# Empezamos a leer desde la fila 5
fila = 5

# Iteramos mientras haya datos en la columna A
while hoja.cell(row=fila, column=1).value:
    # Extraemos los valores de las columnas A, C, D, E, F, G, I, J
    valores_fila = [hoja.cell(row=fila, column=columna).value for columna in range(1, 11) if columna in [1, 3, 4, 5, 6, 7, 9, 10]]
        
    # Multiplicamos el valor de la columna D por 7/6
    valores_fila[2] = valores_fila[2] * 7/6

    # Agregamos los valores a la lista de datos
    datos_de_captura.append(valores_fila)

    # Avanzamos a la siguiente fila
    fila += 1


# GENERAMOS LA LISTA DE ACTIVIDADES
# Creamos una sublista que representa las actividades de cada trabajador
actividades = []
for sublista in datos_de_captura:
    actividad = sublista[5]
    if actividad is None or actividad == "":
        actividad = "" # Si actividad es None o una cadena vacía, asignamos una cadena vacía
    actividades.append(actividad)

lista_de_actividades = [[] for i in range(len(actividades))]

for i, actividad in enumerate(actividades):
    # Dividir la cadena de texto en subcadenas de máximo 46 caracteres
    subcadenas = []
    while len(actividad) > 0:
        if len(actividad) <= 46:
            subcadenas.append(actividad)
            actividad = ""
        else:
            espacio = actividad.rfind(" ", 0, 46)
            if espacio == -1:
                subcadenas.append(actividad[:46])
                actividad = actividad[46:]
            else:
                subcadenas.append(actividad[:espacio])
                actividad = actividad[espacio+1:]
    
    # Agregar las subcadenas a la nueva lista correspondiente
    lista_de_actividades[i].extend(subcadenas)

# GENERAMOS LA LISTA DE TRABAJADORES
# La lista trabajadores, contendra las claves de cada uno de los trabajadores en datos_de_captura
trabajadores = [lista[0] for lista in datos_de_captura]

# Este ciclo for llenara la lista trabajador, enumerando a trabajadores iniciando desde el 0, para poder usarla como parametro en nuestra variable
# Enumeramos los elementos de la lista y guardamos los índices en una lista
trabajador = [i for i, _ in enumerate(trabajadores)]

# FUNCION PARA CAPTURAR EN EL ARCHIVO EXCEL: 
""" Esta funcion realiza una búsqueda en un archivo de Excel que contiene un registro de reporte personal de un trabajador específico. El objetivo es buscar el último valor numérico menor a 70 en la columna B, guardar la coordenada de la celda y devolverla en la lista datos_capturados. En caso de no encontrar ningún valor numérico menor a 70, se devolverá la coordenada de la celda A15.

Después de encontrar la coordenada de la celda, la función realiza varias tareas adicionales, como cargar el archivo de Excel, asignar valores a celdas específicas y guardar el archivo con los cambios. """

def capturar_reporte_personal(trabajador):

    datos_capturados = []
    clave_obra_destino = 1

    # obtener el valor de búsqueda
    obra = datos_de_captura[trabajador][clave_obra_destino]

    # obtener la ruta de búsqueda
    ruta_busqueda = f'/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/SEMANA_0{num_semana}'

    # buscar archivos en la ruta de búsqueda que inicien con el valor de búsqueda
    for archivo in os.listdir(ruta_busqueda):
        if archivo.startswith(obra):
            # si se encuentra el archivo, regresar la ruta completa
            archivo_para_captura = os.path.join(ruta_busqueda, archivo)
            datos_capturados.append(archivo_para_captura)
            break
    
     # si no se encontró ningún archivo, regresar None
    if archivo_para_captura is None:
        print(f"No se encontro la clave {obra} para el trabajador no: {trabajador}")
        return None

    # Cargamos el archivo_para_captura de Excel
    wb = openpyxl.load_workbook(archivo_para_captura)
    # Seleccionamos la hoja en la que queremos buscar
    ws = wb.active

    # Inicializamos las variables para tener el ultimo valor de orden o la ultima celda con texto
    ult_celda_con_valor = None
    ult_texto = None
    ult_celda_con_texto = None

    # Ciclo for para conseguir la ultima celda con valor
    for fila in range(14, 301):
        # Obtenemos el valor de la celda B en la fila actual
        valor_celda = ws.cell(row=fila, column=2).value
        # Si el valor es un número menor a 70, lo almacenamos
        if isinstance(valor_celda, (int, float)) and valor_celda < 70:
            ult_celda_con_valor = ws.cell(row=fila, column=2).coordinate
    ult_celda_con_valor = ult_celda_con_valor
    
    # Ciclo for para conseguir la ultima celda con valor
    # Recorremos las filas desde la 13 hasta la 301
    for fila in range(14, 301):
        # Obtenemos el valor de la celda D en la fila actual
        valor_celda = ws.cell(row=fila, column=4).value
        # Si el valor es un string, lo almacenamos
        if isinstance(valor_celda, str):
            ult_texto = valor_celda
            ult_celda_con_texto = ws.cell(row=fila, column=4).coordinate
    ult_celda_con_texto = ult_celda_con_texto
    ult_texto = ult_texto

    # Logica para asignar las celdas de captura dependiendo el ultimo valor
    if ult_celda_con_valor == None:
        nueva_fila = 15
        nueva_columna = 1
        celda_para_captura = ws.cell(row=nueva_fila, column=nueva_columna).coordinate
    
    elif ult_celda_con_valor is not None:
        if ult_texto is None:
            celda = ws[ult_celda_con_valor]
            nueva_fila = celda.row +1
            nueva_columna = celda.column -1
            celda_para_captura = ws.cell(row=nueva_fila, column=nueva_columna).coordinate
        elif ult_texto.startswith("Tiempo extra"):
            celda = ws[ult_celda_con_texto]
            nueva_fila = celda.row +1
            nueva_columna = celda.column -3
            celda_para_captura = ws.cell(row=nueva_fila, column=nueva_columna).coordinate
        elif ult_texto == "Domingo trabajado":
            celda = ws[ult_celda_con_texto]
            nueva_fila = celda.row +1
            nueva_columna = celda.column -3
            celda_para_captura = ws.cell(row=nueva_fila, column=nueva_columna).coordinate
        else:
            celda = ws[ult_celda_con_texto]
            nueva_fila = celda.row +2
            nueva_columna = celda.column -3
            celda_para_captura = ws.cell(row=nueva_fila, column=nueva_columna).coordinate

        
    return celda_para_captura

res1 = capturar_reporte_personal(1)
print(res1)




    
