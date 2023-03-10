# LIBRERIAS
import openpyxl
import os

# VARIABLES GLOBALES
num_semana = 8
ruta_proyecto = '/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/'
nombre_archivo = f'SEMANA_0{num_semana}_REPORTE.xlsx'

ruta_archivo_origen = os.path.join(ruta_proyecto, f'SEMANA_0{num_semana}', nombre_archivo)
ruta_archivo_destino = os.path.join(ruta_proyecto, f'SEMANA_0{num_semana}')

def obtener_datos_de_captura(ruta_archivo_origen):
    
    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen)
    hoja = libro.active

    # Definimos la lista donde vamos a guardar los datos
    datos_de_captura = []

    # Empezamos a leer desde la fila 5
    fila = 5

    # Iteramos mientras haya datos en la columna A
    while hoja.cell(row=fila, column=1).value:
        # Extraemos los valores de las columnas A, C, D, E, F, G, I, J
        valores_fila = [hoja.cell(row=fila, column=columna).value for columna in range(1, 11) if columna in [1, 3, 4, 5, 6, 7, 9, 10]]

        # Multiplicamos el valor de la columna D por 7/6
        valores_fila[2] = valores_fila[2] * 7/6

        # Agregamos los valores a la lista de datos
        datos_de_captura.append(valores_fila)

        # Avanzamos a la siguiente fila
        fila += 1

    # GENERAMOS LA LISTA DE ACTIVIDADES
    # Creamos una sublista que representa las actividades de cada trabajador
    actividades = []
    for sublista in datos_de_captura:
        actividad = sublista[5]
        if actividad is None or actividad == "":
            actividad = "" # Si actividad es None o una cadena vacía, asignamos una cadena vacía
        actividades.append(actividad)

    lista_de_actividades = [[] for i in range(len(actividades))]

    for i, actividad in enumerate(actividades):
        # Dividir la cadena de texto en subcadenas de máximo 46 caracteres
        subcadenas = []
        while len(actividad) > 0:
            if len(actividad) <= 46:
                subcadenas.append(actividad)
                actividad = ""
            else:
                espacio = actividad.rfind(" ", 0, 46)
                if espacio == -1:
                    subcadenas.append(actividad[:46])
                    actividad = actividad[46:]
                else:
                    subcadenas.append(actividad[:espacio])
                    actividad = actividad[espacio+1:]

        # Agregar las subcadenas a la nueva lista correspondiente
        lista_de_actividades[i].extend(subcadenas)

    # GENERAMOS LA LISTA DE TRABAJADORES
    # La lista trabajadores, contendra las claves de cada uno de los trabajadores en datos_de_captura
    trabajadores = [lista[0] for lista in datos_de_captura]

    # Este ciclo for llenara la lista trabajador, enumerando a trabajadores iniciando desde el 0, para poder usarla como parametro en nuestra variable
    # Enumeramos los elementos de la lista y guardamos los índices en una lista
    trabajador = [i for i, _ in enumerate(trabajadores)]

    return datos_de_captura, lista_de_actividades, trabajador

datos_de_captura, lista_de_actividades, trabajador = obtener_datos_de_captura(ruta_archivo_origen)

""" CICLO DE FUNCIONES PRINCIPALES PARA CAPTURA """

for clave in trabajador:

    def obtener_archivo_para_captura(clave):
        clave_de_obra = datos_de_captura[clave][1]
        # obtener la ruta de búsqueda
        ruta_busqueda = f'/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/SEMANA_0{num_semana}'

        # buscar archivos en la ruta de búsqueda que inicien con el valor de búsqueda
        for archivo in os.listdir(ruta_busqueda):
            if archivo.startswith(clave_de_obra + " ") and os.path.isfile(os.path.join(ruta_busqueda, archivo)):
                # si se encuentra el archivo, regresar la ruta completa
                archivo_para_captura = os.path.join(ruta_busqueda, archivo)
        return archivo_para_captura

    archivo_para_captura = obtener_archivo_para_captura(clave)

    def obtener_celda_para_captura(archivo_para_captura):

        # Cargamos el archivo_para_captura de Excel
        wb = openpyxl.load_workbook(archivo_para_captura)
        # Seleccionamos la hoja en la que queremos buscar
        ws = wb.active

        # Inicializar las variables que almacenarán la celda con el último valor encontrado
        ultima_fila_b = 0
        ultima_fila_d = 0

        # Recorrer las filas del rango especificado
        for fila in range(14, 301):

            # Obtener el valor de la columna B en la fila actual
            valor_b = ws.cell(row=fila, column=2).value
            # Si el valor es un número menor a 70, lo almacenamos
            if isinstance(valor_b, (int, float)) and valor_b < 70:
                ultima_fila_b = fila

            # Obtener el valor de la columna D en la fila actual
            valor_d = ws.cell(row=fila, column=4).value
            # Si el valor es un string, lo almacenamos
            if isinstance(valor_d, str):
                ultima_fila_d = fila
        
        # Obtener la celda para captura

        # Si la fila de b y d son 0, retornaremos la celda inicial
        if ultima_fila_b == 0 and ultima_fila_d == 0:
            celda_para_captura = ws.cell(row=15, column=1)
        # Si la fila b es igual o mayor que fila b, retornamos celda_captura
        elif ultima_fila_b >= ultima_fila_d:
            celda_para_captura = ws.cell(row=ultima_fila_b +1, column=1)
        else:
            celda_para_captura = ws.cell(row=ultima_fila_d +2, column=1)

        # Obetener el ultimo valor

        # Empezamos a buscar desde la fila 14
        fila_actual = 14
        
        # Inicializamos el valor a devolver con None
        ultimo_valor = None
        
        # Recorremos todas las filas de la hoja hasta encontrar un valor numérico menor a 70
        while fila_actual <= 300:
            celda_b = ws.cell(row=fila_actual, column=2)
            valor_b = celda_b.value
            
            # Si la celda B de la fila actual tiene un valor numérico, lo guardamos como último valor
            if isinstance(valor_b, (int, float)):
                if valor_b < 70 and (ultimo_valor is None or valor_b > ultimo_valor):
                    ultimo_valor = valor_b
            
            fila_actual += 1
        
        # Si no se encontró ningún valor menor a 70, se devuelve 1
        if ultimo_valor is None:
            ultimo_valor = 1
        else:
            ultimo_valor += 1



        return celda_para_captura, ultimo_valor
            
    celda_para_captura, ultimo_valor = obtener_celda_para_captura(archivo_para_captura)

    def capturar_trabajador(clave):

        ultimo_valor
        # Cargamos el archivo_para_captura de Excel
        wb = openpyxl.load_workbook(archivo_para_captura)
        # Seleccionamos la hoja en la que queremos buscar
        ws = wb.active

        # Tomamos el valor de la ultima celda de captura
        fila = celda_para_captura.row
        columna = celda_para_captura.column

        # Asignar el codigo
        codigo = ws.cell(row=fila, column=columna)
        codigo.value = datos_de_captura[clave][0]

        # Asignar los valores:
        if datos_de_captura[clave][3] is None and datos_de_captura[clave][4] is None:

            celda_orden1 = ws.cell(row=fila, column=columna +1)
            celda_orden1.value = ultimo_valor
            celda_orden2 = ws.cell(row=fila, column=columna +15)
            celda_orden2.value = ultimo_valor
            celda_dias = ws.cell(row=fila, column=columna +11)
            celda_dias.value = datos_de_captura[clave][2]
            celda_porcentaje = ws.cell(row=fila, column=columna +18)
            celda_porcentaje.value = 1

        elif datos_de_captura[clave][3] is not None and datos_de_captura[clave][4] is None:
            
            celda_orden1 = ws.cell(row=fila, column=columna +1)
            celda_orden1.value = ultimo_valor
            celda_horas = ws.cell(row=fila +1, column=columna)
            celda_horas.value = "lote"
            celda_orden2 = ws.cell(row=fila +1, column=columna +1)
            celda_orden2.value = ultimo_valor
            celda_horas2 = ws.cell(row=fila +1, column=columna +8)
            celda_horas2.value = float(datos_de_captura[clave][7]) * 0.0025 
            celda_orden3 = ws.cell(row=fila +1, column=columna +15)
            celda_orden3.value = ultimo_valor
            celda_horas3 = ws.cell(row=fila +1, column=columna +11)
            celda_horas3.value = datos_de_captura[clave][3]
            celda_horas4 = ws.cell(row=fila +1, column=columna +3)
            celda_horas4.value = (f"Tiempo extra, {datos_de_captura[clave][3]} horas trabajadas")
            celda_dias = ws.cell(row=fila, column=columna +11)
            celda_dias.value = datos_de_captura[clave][2]
            celda_porcentaje = ws.cell(row=fila, column=columna +18)
            celda_porcentaje.value = 1

        elif datos_de_captura[clave][3] is None and datos_de_captura[clave][4] is not None:
            
            celda_orden1 = ws.cell(row=fila, column=columna +1)
            celda_orden1.value = ultimo_valor
            celda_domingo = ws.cell(row=fila +1, column=columna)
            celda_domingo.value = "lote"
            celda_orden2 = ws.cell(row=fila +1, column=columna +1)
            celda_orden2.value = ultimo_valor
            celda_domingo2 = ws.cell(row=fila +1, column=columna +8)
            celda_domingo2.value = ((datos_de_captura[clave][7]) / 100)
            celda_orden3 = ws.cell(row=fila +1, column=columna +15)
            celda_orden3.value = ultimo_valor
            celda_domingo3 = ws.cell(row=fila +1, column=columna +11)
            celda_domingo3.value = datos_de_captura[clave][4] + 1
            celda_domingo4 = ws.cell(row=fila +1, column=columna +3)
            celda_domingo4.value = "Domingo trabajado"
            celda_dias = ws.cell(row=fila, column=columna +11)
            celda_dias.value = datos_de_captura[clave][2]
            celda_porcentaje = ws.cell(row=fila, column=columna +18)
            celda_porcentaje.value = 1
        else:

            celda_orden1 = ws.cell(row=fila, column=columna +1)
            celda_orden1.value = ultimo_valor
            celda_orden2 = ws.cell(row=fila +1, column=columna +1)
            celda_orden2.value = ultimo_valor
            celda_orden3 = ws.cell(row=fila +2, column=columna +1)
            celda_orden3.value = ultimo_valor
            celda_orden4 = ws.cell(row=fila +2, column=columna +15)
            celda_orden4.value = ultimo_valor
            celda_dias = ws.cell(row=fila, column=columna +11)
            celda_dias.value = datos_de_captura[clave][2]
            celda_porcentaje = ws.cell(row=fila, column=columna +18)
            celda_porcentaje.value = 1
            celda_horas = ws.cell(row=fila +1, column=columna)
            celda_horas.value = "lote"
            celda_horas2 = ws.cell(row=fila +1, column=columna +8)
            celda_horas2.value = float(datos_de_captura[clave][7]) * 0.0025
            celda_horas3 = ws.cell(row=fila +1, column=columna +11)
            celda_horas3.value = datos_de_captura[clave][3]
            celda_horas4 = ws.cell(row=fila +1, column=columna +3)
            celda_horas4.value = (f"Tiempo extra, {datos_de_captura[clave][3]} horas trabajadas")
            celda_domingo = ws.cell(row=fila +2, column=columna)
            celda_domingo.value = "lote"
            celda_domingo2 = ws.cell(row=fila +2, column=columna +8)
            celda_domingo2.value = ((datos_de_captura[clave][7]) / 100)
            celda_domingo3 = ws.cell(row=fila +2, column=columna +11)
            celda_domingo3.value = datos_de_captura[clave][4] + 1
            celda_domingo4 = ws.cell(row=fila +2, column=columna +3)
            celda_domingo4.value = "Domingo trabajado"
        
        # Asignar actividades

        if datos_de_captura[clave][5] is not None and datos_de_captura[clave][3] is not None and datos_de_captura[clave][4] is not None:

            for valor in lista_de_actividades[clave]:
                celda_actividades = ws.cell(row=fila +3, column=columna +3)
                celda_actividades.value = valor
                fila += 1

        elif datos_de_captura[clave][5] is not None and datos_de_captura[clave][3] is not None and datos_de_captura[clave][4] is None:

            for valor in lista_de_actividades[clave]:
                celda_actividades = ws.cell(row=fila +2, column=columna +3)
                celda_actividades.value = valor
                fila += 1

        elif datos_de_captura[clave][5] is not None and datos_de_captura[clave][3] is None and datos_de_captura[clave][4] is not None:

            for valor in lista_de_actividades[clave]:
                celda_actividades = ws.cell(row=fila +2, column=columna +3)
                celda_actividades.value = valor
                fila += 1

        elif datos_de_captura[clave][5] is not None and datos_de_captura[clave][3] is None and datos_de_captura[clave][4] is None:

            for valor in lista_de_actividades[clave]:
                celda_actividades = ws.cell(row=fila +1, column=columna +3)
                celda_actividades.value = valor
                fila += 1

        else:
            pass


        wb.save(archivo_para_captura)
        wb.close()
        return print("Se ha capturado en la obra", datos_de_captura[clave][1], "el trabajador numero : ", datos_de_captura[clave][0], "En la celda:", celda_para_captura.coordinate)
    
    capturar_trabajador(clave)

print(f"Se ha terminado la captura del reporte :  SEMANA_0{num_semana} ")