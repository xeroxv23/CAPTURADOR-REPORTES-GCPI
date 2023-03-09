# LIBRERIAS
import openpyxl
import os

# VARIABLES GLOBALES
num_semana = 8
ruta_proyecto = '/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/'
nombre_archivo = f'SEMANA_0{num_semana}_REPORTE.xlsx'

ruta_archivo_origen = os.path.join(ruta_proyecto, f'SEMANA_0{num_semana}', nombre_archivo)
ruta_archivo_destino = os.path.join(ruta_proyecto, f'SEMANA_0{num_semana}')

def obtener_datos_de_captura(ruta_archivo_origen):
    
    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen)
    hoja = libro.active

    # Definimos la lista donde vamos a guardar los datos
    datos_de_captura = []

    # Empezamos a leer desde la fila 5
    fila = 5

    # Iteramos mientras haya datos en la columna A
    while hoja.cell(row=fila, column=1).value:
        # Extraemos los valores de las columnas A, C, D, E, F, G, I, J
        valores_fila = [hoja.cell(row=fila, column=columna).value for columna in range(1, 11) if columna in [1, 3, 4, 5, 6, 7, 9, 10]]

        # Multiplicamos el valor de la columna D por 7/6
        valores_fila[2] = valores_fila[2] * 7/6

        # Agregamos los valores a la lista de datos
        datos_de_captura.append(valores_fila)

        # Avanzamos a la siguiente fila
        fila += 1

    # GENERAMOS LA LISTA DE ACTIVIDADES
    # Creamos una sublista que representa las actividades de cada trabajador
    actividades = []
    for sublista in datos_de_captura:
        actividad = sublista[5]
        if actividad is None or actividad == "":
            actividad = "" # Si actividad es None o una cadena vacía, asignamos una cadena vacía
        actividades.append(actividad)

    lista_de_actividades = [[] for i in range(len(actividades))]

    for i, actividad in enumerate(actividades):
        # Dividir la cadena de texto en subcadenas de máximo 46 caracteres
        subcadenas = []
        while len(actividad) > 0:
            if len(actividad) <= 46:
                subcadenas.append(actividad)
                actividad = ""
            else:
                espacio = actividad.rfind(" ", 0, 46)
                if espacio == -1:
                    subcadenas.append(actividad[:46])
                    actividad = actividad[46:]
                else:
                    subcadenas.append(actividad[:espacio])
                    actividad = actividad[espacio+1:]

        # Agregar las subcadenas a la nueva lista correspondiente
        lista_de_actividades[i].extend(subcadenas)

    # GENERAMOS LA LISTA DE TRABAJADORES
    # La lista trabajadores, contendra las claves de cada uno de los trabajadores en datos_de_captura
    trabajadores = [lista[0] for lista in datos_de_captura]

    # Este ciclo for llenara la lista trabajador, enumerando a trabajadores iniciando desde el 0, para poder usarla como parametro en nuestra variable
    # Enumeramos los elementos de la lista y guardamos los índices en una lista
    trabajador = [i for i, _ in enumerate(trabajadores)]

    return datos_de_captura, lista_de_actividades, trabajador

resultado = obtener_datos_de_captura(ruta_archivo_origen)
datos_de_captura = resultado[0]
lista_de_actividades = resultado[1]
trabajador = resultado[2]

def obtener_archivo_para_captura(trabajador):
    clave_de_obra = datos_de_captura[trabajador][1]
    # obtener la ruta de búsqueda
    ruta_busqueda = f'/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/SEMANA_0{num_semana}'

    # buscar archivos en la ruta de búsqueda que inicien con el valor de búsqueda
    for archivo in os.listdir(ruta_busqueda):
        if archivo.startswith(clave_de_obra + " ") and os.path.isfile(os.path.join(ruta_busqueda, archivo)):
            # si se encuentra el archivo, regresar la ruta completa
            archivo_para_captura = os.path.join(ruta_busqueda, archivo)
    return archivo_para_captura

def obtener_celda_de_captura(archivo_para_captura):
    # Cargamos el archivo_para_captura de Excel
    wb = openpyxl.load_workbook(archivo_para_captura)
    # Seleccionamos la hoja en la que queremos buscar
    ws = wb.active

    # Inicializar las variables que almacenarán la celda con el último valor encontrado
    ultima_celda_b = None
    ultima_celda_d = None

    # Recorrer las filas del rango especificado
    for fila in range(14, 301):

        # Obtener el valor de la columna B en la fila actual
        valor_b = ws.cell(row=fila, column=2).value
        # Si el valor es un número menor a 70, lo almacenamos
        if isinstance(valor_b, (int, float)) and valor_b < 70:
            ultima_celda_b = ws.cell(row=fila, column=2)

        # Obtener el valor de la columna D en la fila actual
        valor_d = ws.cell(row=fila, column=4).value
        # Si el valor es un string, lo almacenamos
        if isinstance(valor_d, str):
            ultima_celda_d = ws.cell(row=fila, column=4)
        
        # Obtener la celda para captura
        if ultima_celda_b is not None and ultima_celda_d is not None:
            # Si se encontró una celda en ambas columnas, seleccionar la que tenga el row mayor
            ultima_celda = ultima_celda_b if ultima_celda_b.row > ultima_celda_d.row else ultima_celda_d
        elif ultima_celda_b is not None:
            # Si solo se encontró una celda en la columna B, usar esa celda
            ultima_celda = ultima_celda_b
        elif ultima_celda_d is not None:
            # Si solo se encontró una celda en la columna D, usar esa celda
            ultima_celda = ultima_celda_d
        else:
            # Si no se encontró ninguna celda, seleccionar la celda en la fila 15, columna 1
            ultima_celda = ws.cell(row=15, column=1)
        
        # Reemplazar a que siempre sea la columna A
        ultima_celda.column = 1
       # Obtener el número de fila y columna de la celda
        fila_actual = ultima_celda.row
        columna_actual = ultima_celda.column
        # Sumar 1 al número de fila
        nueva_fila = fila_actual
        # Crear una nueva instancia de la clase Cell con la misma columna y la nueva fila
        celda_para_captura = ultima_celda.parent.cell(row=nueva_fila, column=columna_actual)

    celda_para_captura

    # Empezamos a buscar desde la fila 14
    fila_actual = 14
    
    # Inicializamos el valor a devolver con None
    ultimo_valor = None
    
    # Recorremos todas las filas de la hoja hasta encontrar un valor numérico menor a 70
    while fila_actual <= 300:
        celda_b = ws.cell(row=fila_actual, column=2)
        valor_b = celda_b.value
        
        # Si la celda B de la fila actual tiene un valor numérico, lo guardamos como último valor
        if isinstance(valor_b, (int, float)):
            if valor_b < 70 and (ultimo_valor is None or valor_b > ultimo_valor):
                ultimo_valor = valor_b
        
        fila_actual += 1
    
    # Si no se encontró ningún valor menor a 70, se devuelve 1
    if ultimo_valor is None:
        ultimo_valor = 1
    ultimo_valor

    return celda_para_captura , ultimo_valor

resultado2 = obtener_celda_de_captura(trabajador)
celda_para_captura = resultado2[0]
ultimo_valor = resultado2[1]


def capturador_de_datos(trabajador):


    # Cargamos el archivo_para_captura de Excel
    wb = openpyxl.load_workbook(archivo_para_captura)
    # Seleccionamos la hoja en la que queremos buscar
    ws = wb.active

    # Crear la celda del codigo y asignarle la clave del trabajador
    fila_codigo = celda_para_captura.row
    columna_codigo = celda_para_captura.column
    celda_codigo = celda_para_captura.parent.cell(row=fila_codigo, column=columna_codigo)
    celda_codigo.value = datos_de_captura[trabajador][0]

    # Asignar los valores:
    if datos_de_captura[trabajador][3] is None and datos_de_captura[trabajador][4] is None:
        fila_orden = celda_para_captura.row
        columna_orden = celda_para_captura.column
        celda_orden1 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +1)
        celda_orden1.value = ultimo_valor
        celda_orden2 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +15)
        celda_orden2.value = ultimo_valor
        celda_dias = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +11)
        celda_dias.value = datos_de_captura[trabajador][2]
        celda_porcentaje = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +18)
        celda_porcentaje.value = 1

    elif datos_de_captura[trabajador][3] is not None and datos_de_captura[trabajador][4] is None:
        fila_orden = celda_para_captura.row
        columna_orden = celda_para_captura.column
        celda_orden1 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +1)
        celda_orden1.value = ultimo_valor
        celda_horas = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden)
        celda_horas.value = "lote"
        celda_orden2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +1)
        celda_orden2.value = ultimo_valor
        celda_horas2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +8)
        celda_horas2.value = float(datos_de_captura[trabajador][7]) * 0.0025 
        celda_orden3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +15)
        celda_orden3.value = ultimo_valor
        celda_horas3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +11)
        celda_horas3.value = datos_de_captura[trabajador][3]
        celda_horas4 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +3)
        celda_horas4.value = (f"Tiempo extra, {datos_de_captura[trabajador][3]} horas trabajadas")
        celda_dias = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +11)
        celda_dias.value = datos_de_captura[trabajador][2]
        celda_porcentaje = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +18)
        celda_porcentaje.value = 1

    elif datos_de_captura[trabajador][3] is None and datos_de_captura[trabajador][4] is not None:
        fila_orden = celda_para_captura.row
        columna_orden = celda_para_captura.column
        celda_orden1 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +1)
        celda_orden1.value = ultimo_valor
        celda_domingo = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden)
        celda_domingo.value = "lote"
        celda_orden2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +1)
        celda_orden2.value = ultimo_valor
        celda_domingo2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +8)
        celda_domingo2.value = ((datos_de_captura[trabajador][7]) / 100)
        celda_orden3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +15)
        celda_orden3.value = ultimo_valor
        celda_domingo3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +11)
        celda_domingo3.value = datos_de_captura[trabajador][4] + 1
        celda_domingo4 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +3)
        celda_domingo4.value = "Domingo trabajado"
        celda_dias = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +11)
        celda_dias.value = datos_de_captura[trabajador][2]
        celda_porcentaje = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +18)
        celda_porcentaje.value = 1
    else:
        fila_orden = celda_para_captura.row
        columna_orden = celda_para_captura.column
        celda_orden1 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +1)
        celda_orden1.value = ultimo_valor
        celda_orden2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +1)
        celda_orden2.value = ultimo_valor
        celda_orden3 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +1)
        celda_orden3.value = ultimo_valor
        celda_orden4 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +15)
        celda_orden4.value = ultimo_valor
        celda_dias = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +11)
        celda_dias.value = datos_de_captura[trabajador][2]
        celda_porcentaje = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +18)
        celda_porcentaje.value = 1
        celda_horas = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden)
        celda_horas.value = "lote"
        celda_horas2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +8)
        celda_horas2.value = float(datos_de_captura[trabajador][7]) * 0.0025
        celda_horas3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +11)
        celda_horas3.value = datos_de_captura[trabajador][3]
        celda_horas4 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +3)
        celda_horas4.value = (f"Tiempo extra, {datos_de_captura[trabajador][3]} horas trabajadas")
        celda_domingo = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden)
        celda_domingo.value = "lote"
        celda_domingo2 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +8)
        celda_domingo2.value = ((datos_de_captura[trabajador][7]) / 100)
        celda_domingo3 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +11)
        celda_domingo3.value = datos_de_captura[trabajador][4] + 1
        celda_domingo4 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +3)
        celda_domingo4.value = "Domingo trabajado"

    # Asignar actividades

    if datos_de_captura[trabajador][5] is not None and datos_de_captura[trabajador][3] is not None and datos_de_captura[trabajador][4] is not None:

        for valor in lista_de_actividades[trabajador]:
            celda_actividades = celda_para_captura.parent.cell(row=fila_orden +3, column=columna_orden +3)
            celda_actividades.value = valor
            fila_orden += 1

    elif datos_de_captura[trabajador][5] is not None and datos_de_captura[trabajador][3] is not None and datos_de_captura[trabajador][4] is None:

        for valor in lista_de_actividades[trabajador]:
            celda_actividades = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +3)
            celda_actividades.value = valor
            fila_orden += 1

    elif datos_de_captura[trabajador][5] is not None and datos_de_captura[trabajador][3] is None and datos_de_captura[trabajador][4] is not None:

        for valor in lista_de_actividades[trabajador]:
            celda_actividades = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +3)
            celda_actividades.value = valor
            fila_orden += 1

    elif datos_de_captura[trabajador][5] is not None and datos_de_captura[trabajador][3] is None and datos_de_captura[trabajador][4] is None:

        for valor in lista_de_actividades[trabajador]:
            celda_actividades = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +3)
            celda_actividades.value = valor
            fila_orden += 1

    else:
        pass


    wb.save(archivo_para_captura)
    wb.close()

    return print("Se ha capturado en la obra", datos_de_captura[trabajador][1], "el trabajador numero : ", datos_de_captura[trabajador][0], celda_para_captura.coordinate)


# El ciclo for que capturara todos los valores en los archivos de excel, recorriendo en el parametro de la funcion cada uno de los valores del reporte
for clave in trabajador:
    archivo_para_captura = obtener_archivo_para_captura(clave)
    celda_para_captura = obtener_celda_de_captura(archivo_para_captura)
    capturador_de_datos(clave)
print("Se capturaron los trabajadores") 