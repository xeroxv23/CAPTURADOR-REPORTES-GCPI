import openpyxl
import os
import re
from capturador import datos_de_captura

num_semana = 8

def prueba_captura(trabajador):
    clave_de_obra = datos_de_captura[trabajador][1]
    # obtener la ruta de búsqueda
    ruta_busqueda = f'/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/SEMANA_0{num_semana}'

    # buscar archivos en la ruta de búsqueda que inicien con el valor de búsqueda
    for archivo in os.listdir(ruta_busqueda):
        if archivo.startswith(clave_de_obra + " ") and os.path.isfile(os.path.join(ruta_busqueda, archivo)):
            # si se encuentra el archivo, regresar la ruta completa
            archivo_para_captura = os.path.join(ruta_busqueda, archivo)
    
    # Cargamos el archivo_para_captura de Excel
    wb = openpyxl.load_workbook(archivo_para_captura)
    # Seleccionamos la hoja en la que queremos buscar
    ws = wb.active

    # Inicializar las variables que almacenarán la celda con el último valor encontrado
    ultima_celda_b = None
    ultima_celda_d = None

    # Recorrer las filas del rango especificado
    for fila in range(14, 301):

        # Obtener el valor de la columna B en la fila actual
        valor_b = ws.cell(row=fila, column=2).value
        # Si el valor es un número menor a 70, lo almacenamos
        if isinstance(valor_b, (int, float)) and valor_b < 70:
            ultima_celda_b = ws.cell(row=fila, column=2)

        # Obtener el valor de la columna D en la fila actual
        valor_d = ws.cell(row=fila, column=4).value
        # Si el valor es un string, lo almacenamos
        if isinstance(valor_d, str):
            ultima_celda_d = ws.cell(row=fila, column=4)
        
        # Obtener la celda para captura
        if ultima_celda_b is not None and ultima_celda_d is not None:
            # Si se encontró una celda en ambas columnas, seleccionar la que tenga el row mayor
            ultima_celda = ultima_celda_b if ultima_celda_b.row > ultima_celda_d.row else ultima_celda_d
        elif ultima_celda_b is not None:
            # Si solo se encontró una celda en la columna B, usar esa celda
            ultima_celda = ultima_celda_b
        elif ultima_celda_d is not None:
            # Si solo se encontró una celda en la columna D, usar esa celda
            ultima_celda = ultima_celda_d
        else:
            # Si no se encontró ninguna celda, seleccionar la celda en la fila 15, columna 1
            ultima_celda = ws.cell(row=15, column=1)
        
        # Reemplazar a que siempre sea la columna A
        ultima_celda.column = 1
       # Obtener el número de fila y columna de la celda
        fila_actual = ultima_celda.row
        columna_actual = ultima_celda.column
        # Sumar 1 al número de fila
        nueva_fila = fila_actual
        # Crear una nueva instancia de la clase Cell con la misma columna y la nueva fila
        celda_para_captura = ultima_celda.parent.cell(row=nueva_fila, column=columna_actual)

    celda_para_captura

    # Empezamos a buscar desde la fila 14
    fila_actual = 14
    
    # Inicializamos el valor a devolver con None
    ultimo_valor = None
    
    # Recorremos todas las filas de la hoja hasta encontrar un valor numérico menor a 70
    while fila_actual <= 300:
        celda_b = ws.cell(row=fila_actual, column=2)
        valor_b = celda_b.value
        
        # Si la celda B de la fila actual tiene un valor numérico, lo guardamos como último valor
        if isinstance(valor_b, (int, float)):
            if valor_b < 70 and (ultimo_valor is None or valor_b > ultimo_valor):
                ultimo_valor = valor_b
        
        fila_actual += 1
    
    # Si no se encontró ningún valor menor a 70, se devuelve 1
    if ultimo_valor is None:
        ultimo_valor = 1
    ultimo_valor

    """ SIGUE EL PROCESO MAS IMPORTANTE DE LA FUNCION Y ES CAPTURAR LOS DATOS DE LA LISTA DEL TRABAJADOR CORRIENDO EN LA FUNCION, EN LAS CELDAS QUE YA FUERON OBTENIDAS"""

    # Crear la celda del codigo y asignarle la clave del trabajador
    fila_codigo = celda_para_captura.row
    columna_codigo = celda_para_captura.column
    celda_codigo = celda_para_captura.parent.cell(row=fila_codigo, column=columna_codigo)
    celda_codigo.value = datos_de_captura[trabajador][0]

    # Asignar los valores:
    if datos_de_captura[trabajador][3] is None and datos_de_captura[trabajador][4] is None:
        fila_orden = celda_para_captura.row
        columna_orden = celda_para_captura.column
        celda_orden1 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +1)
        celda_orden1.value = ultimo_valor
        celda_orden2 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +15)
        celda_orden2.value = ultimo_valor
        celda_dias = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +11)
        celda_dias.value = datos_de_captura[trabajador][2]
        celda_porcentaje = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +18)
        celda_porcentaje.value = 1

    elif datos_de_captura[trabajador][3] is not None and datos_de_captura[trabajador][4] is None:
        fila_orden = celda_para_captura.row
        columna_orden = celda_para_captura.column
        celda_orden1 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +1)
        celda_orden1.value = ultimo_valor
        celda_horas = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden)
        celda_horas.value = "lote"
        celda_orden2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +1)
        celda_orden2.value = ultimo_valor
        celda_horas2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +8)
        celda_horas2.value = float(datos_de_captura[trabajador][7]) * 0.0025 
        celda_orden3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +15)
        celda_orden3.value = ultimo_valor
        celda_horas3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +11)
        celda_horas3.value = datos_de_captura[trabajador][3]
        celda_horas4 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +3)
        celda_horas4.value = (f"Tiempo extra, {datos_de_captura[trabajador][3]} horas trabajadas")
        celda_dias = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +11)
        celda_dias.value = datos_de_captura[trabajador][2]
        celda_porcentaje = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +18)
        celda_porcentaje.value = 1

    elif datos_de_captura[trabajador][3] is None and datos_de_captura[trabajador][4] is not None:
        fila_orden = celda_para_captura.row
        columna_orden = celda_para_captura.column
        celda_orden1 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +1)
        celda_orden1.value = ultimo_valor
        celda_domingo = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden)
        celda_domingo.value = "lote"
        celda_orden2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +1)
        celda_orden2.value = ultimo_valor
        celda_domingo2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +8)
        celda_domingo2.value = ((datos_de_captura[trabajador][7]) / 100)
        celda_orden3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +15)
        celda_orden3.value = ultimo_valor
        celda_domingo3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +11)
        celda_domingo3.value = datos_de_captura[trabajador][4] + 1
        celda_domingo4 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +3)
        celda_domingo4.value = "Domingo trabajado"
        celda_dias = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +11)
        celda_dias.value = datos_de_captura[trabajador][2]
        celda_porcentaje = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +18)
        celda_porcentaje.value = 1
    else:
        fila_orden = celda_para_captura.row
        columna_orden = celda_para_captura.column
        celda_orden1 = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +1)
        celda_orden1.value = ultimo_valor
        celda_orden2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +1)
        celda_orden2.value = ultimo_valor
        celda_orden3 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +1)
        celda_orden3.value = ultimo_valor
        celda_orden4 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +15)
        celda_orden4.value = ultimo_valor
        celda_dias = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +11)
        celda_dias.value = datos_de_captura[trabajador][2]
        celda_porcentaje = celda_para_captura.parent.cell(row=fila_orden, column=columna_orden +18)
        celda_porcentaje.value = 1
        celda_horas = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden)
        celda_horas.value = "lote"
        celda_horas2 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +8)
        celda_horas2.value = float(datos_de_captura[trabajador][7]) * 0.0025
        celda_horas3 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +11)
        celda_horas3.value = datos_de_captura[trabajador][3]
        celda_horas4 = celda_para_captura.parent.cell(row=fila_orden +1, column=columna_orden +3)
        celda_horas4.value = (f"Tiempo extra, {datos_de_captura[trabajador][3]} horas trabajadas")
        celda_domingo = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden)
        celda_domingo.value = "lote"
        celda_domingo2 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +8)
        celda_domingo2.value = ((datos_de_captura[trabajador][7]) / 100)
        celda_domingo3 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +11)
        celda_domingo3.value = datos_de_captura[trabajador][4] + 1
        celda_domingo4 = celda_para_captura.parent.cell(row=fila_orden +2, column=columna_orden +3)
        celda_domingo4.value = "Domingo trabajado"

        

    wb.save(archivo_para_captura)
    return print("Se capturo el trabajador")

prueba_captura(1)


"""return archivo_para_captura, ultima_celda_b, ultima_celda_d, celda_para_captura, ultimo_valor
        
# Llamar a la funcion
archivo_para_captura, ultima_celda_b, ultima_celda_d, celda_para_captura, ultimo_valor  = prueba_captura(0)

# Imprimir las coordenadas de las últimas celdas encontradas
print(celda_para_captura)"""

