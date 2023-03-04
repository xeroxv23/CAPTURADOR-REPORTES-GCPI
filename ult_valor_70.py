import openpyxl


# Cargamos el archivo_para_captura de Excel
    wb = openpyxl.load_workbook(archivo_para_captura)
    # Seleccionamos la hoja en la que queremos buscar
    ws = wb.active
    # Inicializamos las variables para almacenar la última celda con un valor menor a 70
    ult_valor_menor_70 = None
    ult_celda_con_valor = None
    # Recorremos las filas desde la 14 hasta la 300
    for fila in range(14, 301):
        # Obtenemos el valor de la celda B en la fila actual
        valor_celda = ws.cell(row=fila, column=2).value
        # Si el valor es un número menor a 70, lo almacenamos
        if isinstance(valor_celda, (int, float)) and valor_celda < 70:
            ult_valor_menor_70 = valor_celda
            ult_celda_con_valor = ws.cell(row=fila, column=2).coordinate
    # Si encontramos un valor menor a 70, retornamos la coordenada de la última celda encontrada
    if ult_celda_con_valor is not None:
        celda = ws[ult_celda_con_valor]
        nueva_fila = celda.row + 3
        nueva_columna = celda.column - 1
        celda_para_captura = ws.cell(row=nueva_fila, column=nueva_columna)

        datos_capturados.append(celda_para_captura.coordinate)
    # Si no encontramos ningún valor menor a 70, retornamos la celda A15
    else:
        nueva_fila = 15
        nueva_columna = 1
        celda_para_captura = ws.cell(row=nueva_fila, column=nueva_columna).coordinate
        datos_capturados.append(celda_para_captura)