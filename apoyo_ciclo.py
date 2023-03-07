import openpyxl

# Cargamos el archivo de Excel
wb = openpyxl.load_workbook('/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/SEMANA_08/A-002 Mantenimientos Av España.xlsm')

# Seleccionamos la hoja en la que queremos buscar
ws = wb.active

# Inicializamos las variables para almacenar la última celda con texto
ult_texto = None
ult_celda_con_texto = None

# Recorremos las filas desde la 13 hasta la 301
for fila in range(14, 301):
    # Obtenemos el valor de la celda D en la fila actual
    valor_celda = ws.cell(row=fila, column=4).value
    # Si el valor es un string, lo almacenamos
    if isinstance(valor_celda, str):
        ult_texto = valor_celda
        ult_celda_con_texto = ws.cell(row=fila, column=4).coordinate
    
# Inicializamos las variables para almacenar la última celda con un valor menor a 70
ult_celda_con_valor = None

# Recorremos las filas desde la 14 hasta la 300
for fila in range(14, 301):
    # Obtenemos el valor de la celda B en la fila actual
    valor_celda = ws.cell(row=fila, column=2).value
    # Si el valor es un número menor a 70, lo almacenamos
    if isinstance(valor_celda, (int, float)) and valor_celda < 70:
        ult_celda_con_valor = ws.cell(row=fila, column=2).coordinate

# Mostramos los resultados
print("La última cadena de caracteres encontrada es:", ult_texto)
print("Su celda correspondiente es:", ult_celda_con_texto)
print("La ultima celda con valor encontrada es:", ult_celda_con_valor)