# LIBRERIAS
import openpyxl
import pandas as pd

# VARIABLES GLOBALES
num_semana = 8
ruta_archivo_origen = f'/home/xeroxv23/Documents/Proyectos GCPI/reportes_personal_zonaindustrial/SEMANA_0{num_semana}/SEMANA_0{num_semana}_REPORTE.xlsx'

def captura_reporte(num_semana):
    
    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen)
    hoja = libro.active

    # Definimos la lista donde vamos a guardar los datos
    datos_de_captura = []

    # Empezamos a leer desde la fila 5
    fila = 5

    # Iteramos mientras haya datos en la columna A
    while hoja.cell(row=fila, column=1).value:
        # Extraemos los valores de las columnas A, C, D, E, F, G, H
        valores_fila = [hoja.cell(row=fila, column=columna).value for columna in range(1, 9) if columna in [1, 3, 4, 5, 6, 7, 8]]
        
        # Agregamos los valores a la lista de datos
        datos_de_captura.append(valores_fila)

        # Avanzamos a la siguiente fila
        fila += 1

    # Devolvemos la lista de datos
    return datos_de_captura

datos_de_captura = captura_reporte(8)
print(captura_reporte(8))


def extraer_lista_claves(datos_de_captura):

    # Creamos una lista vacía para guardar los segundos valores
    lista_de_claves = []

    # Iteramos sobre las tuplas de la lista de datos_de_captura
    for tupla in datos_de_captura:
        # Añadimos el segundo valor de cada tupla a la lista de segundos valores
        lista_de_claves.append(tupla[1])

    # Devolvemos la lista de segundos valores
    return lista_de_claves

lista_claves = extraer_lista_claves(datos_de_captura)
print(extraer_lista_claves(datos_de_captura))

    